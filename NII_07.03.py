import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import openpyxl
from glob import glob
import os

lst_nazva = []


def union(units, business, NII, x, y):
    pd.set_option('display.max_rows', 5000)
    pd.set_option('display.max_columns', 5000)
    pd.set_option('display.width', 5000)

    # Заголовочная часть таблицы
    ostatok_1_units = units[['№ (рівень 1)', 'Назва (рівень 1)', '№ (рівень 2)', 'Назва (рівень 2)', '№ (рівень 3)',
                             'Назва (рівень 3)', 'Група статей (рівень.4)',
                             'Наименование группы 4', 'БАНК (п.), грн.', 'БАНК (ф.), грн.', 'БАНК (%)']]

    # Значения с таблицы бизнесов
    business.drop(
        ['№ (рівень 1)', 'Назва (рівень 1)', '№ (рівень 2)', 'Назва (рівень 2)', '№ (рівень 3)', 'Назва (рівень 3)',
         'Група статей (рівень.4)',
         'Наименование группы 4', 'БАНК (п.), грн.', 'БАНК (ф.), грн.', 'БАНК (%)'],
        axis=1,  # удаление строк происходит аналогично, чтобы удалить именно колонки, выбираем соответствующую ось
        inplace=True  # удаление "на месте", без присваивания новой переменной
    )

    bus_copy = business.copy()

    # Выравнивание значений и замена на пустоту

    for i in range(21, y):
        bus_copy.loc[i] = np.nan
    nan = bus_copy.drop(bus_copy.index[0:21])
    nan = nan.drop(nan.index[x:len(nan)])
    df = pd.DataFrame(business[21:len(business)])
    df.index = (df.index + x)
    buss = business.drop(business.index[21:len(business)])
    business1 = pd.concat([buss, nan], axis=0)
    business2 = pd.concat([business1, df], axis=0)



    # Значения с таблицы юнитов
    units.drop(
        ['№ (рівень 1)', 'Назва (рівень 1)', '№ (рівень 2)', 'Назва (рівень 2)', '№ (рівень 3)', 'Назва (рівень 3)',
         'Група статей (рівень.4)',
         'Наименование группы 4', 'БАНК (п.), грн.', 'БАНК (ф.), грн.', 'БАНК (%)'],
        axis=1,  # удаление строк происходит аналогично, чтобы удалить именно колонки, выбираем соответствующую ось
        inplace=True  # удаление "на месте", без присваивания новой переменной
    )

    combiend_df_1 = pd.concat([ostatok_1_units, business2], axis=1)

    # Об'єднана таблиця фул
    combiend_df_2 = pd.concat([combiend_df_1, units], axis=1)

    # ========================================================================================================
    # Таблица для агрегирования
    cd2 = combiend_df_2.copy()
    cd2.drop(
        ['№ (рівень 1)', 'Назва (рівень 1)', '№ (рівень 2)', 'Назва (рівень 2)', '№ (рівень 3)', 'Назва (рівень 3)',
         'Група статей (рівень.4)',
         'Наименование группы 4'],
        axis=1,  # удаление строк происходит аналогично, чтобы удалить именно колонки, выбираем соответствующую ось
        inplace=True  # удаление "на месте", без присваивания новой переменной
    )

    # Агрегирования данных по Комисионным доходам
    combiend_df_2.set_index('№ (рівень 1)', inplace=True)
    riven_4_2 = combiend_df_2.loc['4.2.']

    part1 = riven_4_2.groupby(by=['№ (рівень 1)', 'Назва (рівень 1)'])[cd2.columns].max()
    part2 = riven_4_2.groupby(by=['№ (рівень 2)', 'Назва (рівень 2)'])[cd2.columns].sum()
    part3 = riven_4_2.groupby(by=['№ (рівень 3)', 'Назва (рівень 3)'])[cd2.columns].sum()
    part4 = riven_4_2.groupby(by=['Група статей (рівень.4)', 'Наименование группы 4'])[cd2.columns].sum()

    res = pd.concat([part1, part2, part3, part4], axis=0)

    # Агрегирования данных по Комиссионным затратам
    riven_5_2 = combiend_df_2.loc['5.2.']
    vytraty1 = riven_5_2.groupby(by=['№ (рівень 1)', 'Назва (рівень 1)'])[cd2.columns].min()
    vytraty2 = riven_5_2.groupby(by=['№ (рівень 2)', 'Назва (рівень 2)'])[cd2.columns].sum()
    vytraty3 = riven_5_2.groupby(by=['№ (рівень 3)', 'Назва (рівень 3)'])[cd2.columns].sum()

    res2 = pd.concat([vytraty1, vytraty2, vytraty3], axis=0)

    # Комиссиоонные доходы и затраты
    data = pd.DataFrame(pd.concat([res, res2], axis=0))
    data = data.sort_values(by='№ (рівень 1)')

    # ======================================================================================================================
    # Чистый комиссионный доход
    clear_dohid = data.loc['4.2.', 'Комісійні доходи'] + data.loc['5.2.', 'Комісійні витрати']
    df_cd = pd.DataFrame(clear_dohid).T
    fors = pd.DataFrame(
        data={
            '№ (рівень 1)': [''],
            'Назва (рівень 1)': ['Чистий комісійний дохід']
        })
    clear_dohid_full = pd.concat([fors, df_cd], axis=1)

    def zamena_indexa(table):
        table.reset_index(
            drop=False,  # False означает, что существующий индекс переместится в колонки, а не будет удален
            inplace=True  # произвести операцию на месте, без присваивания новой переменной
        )

    zamena_indexa(data)

    data = pd.concat([data, clear_dohid_full])

    # ======================================================================================================================
    # Чистые доходы
    urovni = combiend_df_2.loc['4.3.':'4.5.']
    r1 = urovni.groupby(by=['№ (рівень 1)', 'Назва (рівень 1)'])[cd2.columns].sum()
    zamena_indexa(r1)

    # Другие доходы
    if '4.6.' in combiend_df_2.iloc[:, 0]:
        urovni2 = combiend_df_2.loc['4.6.':'4.7.']
        r2 = urovni2.groupby(by=['№ (рівень 1)', 'Назва (рівень 1)'])[cd2.columns].sum()
        zamena_indexa(r2)
    else:
        urovni2 = combiend_df_2.loc['4.7.':'4.7.']
        r2 = urovni2.groupby(by=['№ (рівень 1)', 'Назва (рівень 1)'])[cd2.columns].sum()
        zamena_indexa(r2)

    # Другие операционные доходы
    riven_4_8 = combiend_df_2.loc['4.8.']
    insh_op_doh_1 = riven_4_8.groupby(by=['№ (рівень 1)', 'Назва (рівень 1)'])[cd2.columns].max()
    insh_op_doh_2 = riven_4_8.groupby(by=['№ (рівень 2)', 'Назва (рівень 2)'])[cd2.columns].sum()
    res3 = pd.concat([insh_op_doh_1, insh_op_doh_2], axis=0)
    zamena_indexa(res3)

    # Другие доходы
    riven_4_9 = combiend_df_2.loc['4.9.']
    insh_doh_1 = riven_4_9.groupby(by=['№ (рівень 1)', 'Назва (рівень 1)'])[cd2.columns].max()
    insh_doh_2 = riven_4_9.groupby(by=['№ (рівень 2)', 'Назва (рівень 2)'])[cd2.columns].sum()
    res4 = pd.concat([insh_doh_1, insh_doh_2], axis=0)
    zamena_indexa(res4)

    data = pd.concat([data, r1, r2, res3, res4])

    # =======================================================================================================================
    # Внутренние комиссионые доходы и затраты
    vnutrishni = combiend_df_2.loc['4.10.':'5.10.']
    vnutrishni2 = vnutrishni.groupby(by=['№ (рівень 1)', 'Назва (рівень 1)'])[cd2.columns].max()
    vnutrishni2 = vnutrishni2.iloc[:2]

    # Чистые внутренние комиссионые доходы
    clear_vnutri_dohid = vnutrishni2.loc['4.10.', 'Внутрішні комісійні доходи'] + \
                         vnutrishni2.loc['5.10.', 'Внутрішні комісійні витрати']
    clear_vnutri_dohid = pd.DataFrame(clear_vnutri_dohid).T
    fors2 = pd.DataFrame(
        data={
            '№ (рівень 1)': [''],
            'Назва (рівень 1)': ['Чистий внутрішній комісійний дохід']
        })
    clear_vnutri_dohid_full = pd.concat([fors2, clear_vnutri_dohid], axis=1)

    zamena_indexa(vnutrishni2)
    data = pd.concat([data, vnutrishni2, clear_vnutri_dohid_full])

    # Непроцентнные доходы
    neprocentni = combiend_df_2.iloc[-1].drop(['№ (рівень 2)', 'Назва (рівень 2)', '№ (рівень 3)', 'Назва (рівень 3)',
                                               'Група статей (рівень.4)', 'Наименование группы 4'])


    neprocentni[0] = 'Непроцентні доходи'

    neprocentni = pd.DataFrame(neprocentni).T
    data = pd.concat([data, neprocentni])
    data = data.replace(0, np.nan)

    # Расчет процентной части
    data['БАНК (%)'] = round(data['БАНК (ф.), грн.'] / data['БАНК (п.), грн.'] * 100)
    data.iloc[:, 4::3] = np.round(data.iloc[:, 3::3].values / data.iloc[:, 2::3].values * 100)
    data = data.replace(np.nan, 0)
    data.iloc[:, 4::3] = data.iloc[:, 4::3].astype(str) + '%'
    data.iloc[:, 3::3] = (data.iloc[:, 3::3]) / 1000
    data.iloc[:, 2::3] = (data.iloc[:, 2::3]) / 1000


    for i in data.columns[17::3]:
        lst_nazva.append(" ".join(i.split()[:-1]))
        if len(lst_nazva) == 15:
            break

    for i in data.iloc[:, 4::3]:
        data.rename(columns={i: ' % виконання м\п '}, inplace=True)

    for i in data.iloc[:, 3::3]:
        data.rename(columns={i: '      Факт       '}, inplace=True)

    for i in data.iloc[:, 2::3]:
        data.rename(columns={i: '      План       '}, inplace=True)



    data.set_index('№ (рівень 1)', inplace=True)

    data.to_excel(NII)


filename_month_unit = glob('Бізнес-юніти на*.xlsx')[0]
filename_month_business = glob('Бізнес на*.xlsx')[0]

# Извлечение даты с названия файла
date_unit = str(glob('Бізнес-юніти на*.xlsx')[0])
date_unit_full = date_unit.split()[2]
lst_date = []
for i in date_unit_full.split('.'):
    if i.isnumeric() == True:
        lst_date.append(i)
correct_date = "-".join(lst_date)

filename_nakopytel_unit = glob('Бізнес-юніти накопичувально на*.xlsx')[0]
filename_nakopytel_business = glob('Бізнес накопичувально на*.xlsx')[0]

difference_of_cells = int(input("Введите разницу между количеством ячеек входных файлов (Накопичувальний): "))
difference_of_cells_2 = int(input("Введите разницу между количеством ячеек входных файлов (Звичайний): "))

union(pd.read_excel(filename_nakopytel_unit), pd.read_excel(filename_nakopytel_business), 'NII - test year.xlsx', difference_of_cells, difference_of_cells + 21)
union(pd.read_excel(filename_month_unit), pd.read_excel(filename_month_business), 'NII - test month.xlsx', difference_of_cells_2, difference_of_cells_2 + 21)


# =======================================================================================================
# ========================================================================================================
#                                               OPENPYXL


def formating(NII_test, file_NII):
    wb = openpyxl.load_workbook(NII_test)
    Sheet = wb['Sheet1']
    Sheet.title = 'НЕПРОЦЕНТНІ ДОХОДИ'
    ws = wb['НЕПРОЦЕНТНІ ДОХОДИ']

    data = pd.read_excel(NII_test)

    # Формирование списков по урованям
    list_rivni, blue, gray, thin = [], [], [], []
    lst_level = []
    for i in range(2, len(data.values) + 2):
        if ws['A' + str(i)].value == 0:
            ws['A' + str(i)].value = ' '
        list_rivni.append(ws['A' + str(i)].value)
    for i in list_rivni:
        if i is not None:
            lst_level.append(i)
    for i in lst_level:
        if len(str(i)) == 4:
            blue.append(i)
        elif len(str(i)) == 6 or len(str(i)) == 7:
            gray.append(i)
        elif len(str(i)) == 10 or len(str(i)) == 11:
            thin.append(i)

    for column in ws.columns:
        lenght = max(len(str(cell.value)) for cell in column)
        lenght = lenght if lenght <= 75 else 75
        ws.column_dimensions[column[0].column_letter].width = lenght

    col_range = ws.max_column

    for i in range(1, len(data.values) + 2):
        for col in range(1, col_range + 1):
            cell_header = ws.cell(i, col)
            cell_header.number_format = '#,##0'
            cell_header.alignment = Alignment(horizontal='right', vertical='center')
            cell_header.font = Font(name='Times New Roman', size=10, bold=True)
            cell_header.border = Border(left=Side(border_style='thin', color='FF000000'),
                                        right=Side(border_style='thin', color='FF000000'),
                                        top=Side(border_style='thin', color='FF000000'),
                                        bottom=Side(style='thin'))
        ws['B' + str(i)].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws['A' + str(i)].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        names = [ws['B' + str(i)].value]
        indexes = [ws['A' + str(i)].value]
        for j in indexes:
            ws['A' + str(i)].font = Font(name='Times New Roman', size=11, bold=True)
            ws['B' + str(i)].font = Font(name='Times New Roman', size=11, bold=True)
            if j in blue:
                for col in range(1, col_range + 1):
                    cell_header = ws.cell(i, col)
                    cell_header.fill = PatternFill(start_color='B0E0E6', fill_type="solid")
                    cell_header.font = Font(name='Times New Roman', size=12, bold=True)
                    cell_header.number_format = '#,##0'
                ws['B' + str(i)] = ws['B' + str(i)].value.upper()
                ws['B' + str(i)].font = Font(name='Times New Roman', size=12, bold=True)
                ws['A' + str(i)].font = Font(name='Times New Roman', size=12, italic=True, bold=True)
                ws['A' + str(i)].fill = PatternFill(fill_type='solid', start_color='B0E0E6')
                ws['B' + str(i)].fill = PatternFill(fill_type='solid', start_color='B0E0E6')
            elif j in gray:
                for col in range(1, col_range + 1):
                    cell_header = ws.cell(i, col)
                    cell_header.fill = PatternFill(start_color='C0C0C0', fill_type="solid")
                    cell_header.font = Font(name='Times New Roman', size=11, bold=True)
                    cell_header.number_format = '#,##0'
                ws['B' + str(i)] = ws['B' + str(i)].value.upper()
                ws['B' + str(i)].font = Font(name='Times New Roman', size=11, bold=True)
                ws['A' + str(i)].font = Font(name='Times New Roman', size=11, italic=True, bold=True)
                ws['A' + str(i)].fill = PatternFill(fill_type='solid', start_color='DCDCDC')
                ws['B' + str(i)].fill = PatternFill(fill_type='solid', start_color='C0C0C0')
            elif j in thin:
                for col in range(1, col_range + 1):
                    cell_header = ws.cell(i, col)
                    cell_header.font = Font(name='Times New Roman', size=8, italic=True, bold=False)
                    cell_header.number_format = '#,##0'
                ws['B' + str(i)].font = Font(name='Times New Roman', size=8, italic=True, bold=False)
                ws['B' + str(i)].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True, )
                ws['A' + str(i)].font = Font(name='Times New Roman', size=8, italic=True, bold=False)
                # ws.row_dimensions.group(i + 1, hidden=True)

        for k in names:
            if k == 'Чистий комісійний дохід' or k == 'Чистий внутрішній комісійний дохід':
                for col in range(1, col_range + 1):
                    cell_header = ws.cell(i, col)
                    cell_header.fill = PatternFill(start_color='87CEFA', fill_type="solid")
                    cell_header.font = Font(name='Times New Roman', size=11, bold=True)
                    cell_header.number_format = '#,##0'
                ws['A' + str(i)].fill = PatternFill(fill_type='solid', start_color='87CEFA')
                ws['B' + str(i)].font = Font(name='Times New Roman', size=12, bold=True)
                ws['B' + str(i)].fill = PatternFill(fill_type='solid', start_color='87CEFA')


            elif k == 'Непроцентні доходи':
                for col in range(1, col_range + 1):
                    cell_header = ws.cell(i, col)
                    cell_header.fill = PatternFill(start_color='8db4e2', fill_type="solid")
                    cell_header.font = Font(name='Times New Roman', size=12, bold=True)
                    cell_header.number_format = '#,##0'
                ws['B' + str(i)] = ws['B' + str(i)].value.upper()
                ws['A' + str(i)].fill = PatternFill(fill_type='solid', start_color='8db4e2')
                ws['B' + str(i)].font = Font(name='Times New Roman', size=12, bold=True)
                ws['B' + str(i)].fill = PatternFill(fill_type='solid', start_color='8db4e2')
                ws.insert_rows(ws['A' + str(i)].row)


    ws.insert_rows(0)
    ws.insert_rows(0)
    ws.insert_rows(0)
    ws.insert_cols(18)
    column_number = 18
    column = str(chr(64 + column_number))
    ws.column_dimensions[column].width = 3
    for i in range(1, len(data.values) + 4):
        ws['R' + str(i)].fill = PatternFill(fill_type='solid', start_color='B0E0E6')



    list_slice = ['РАЗОМ', 'КОРПОРАТИВНИЙ БІЗНЕС', 'МСБ', 'РОЗДРІБНИЙ БІЗНЕС', 'ІНШІ']
    # for i in lst_nazva[:-1]:
    #     list_slice.append(i)


    ws.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2)
    ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
    ws.merge_cells(start_row=2, start_column=19, end_row=2, end_column=60)
    ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=14)

    for i in range(6, 15):
        ws['S2'] = 'доходи отримані іншими продуктовими підрозділами'
        ws['S2'].font = Font(name='Times New Roman', size=12, bold=True)
        ws['S2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=2, column=i).border = Border(left=Side(border_style='thick', color='FF000000'),
                                    right=Side(border_style='thick', color='FF000000'),
                                    top=Side(border_style='thick', color='FF000000'),
                                    bottom=Side(style='thick'))
    for i in range(3, 18):
        ws.cell(row=3, column=i).border = Border(left=Side(border_style='thick', color='FF000000'),
                                             right=Side(border_style='thick', color='FF000000'),
                                             top=Side(border_style='thick', color='FF000000'),
                                             bottom=Side(style='thick'))


    for i in range(19, 61):
        ws['F2'] = 'в т.ч. бізнес-напрями за сегментами клієнтів відповідного бізнесу'
        ws['F2'].font = Font(name='Times New Roman', size=12, bold=True)
        ws['F2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=2, column=i).border = Border(left=Side(border_style='thick', color='FF000000'),
                                 right=Side(border_style='thick', color='FF000000'),
                                 top=Side(border_style='thick', color='FF000000'),
                                 bottom=Side(style='thick'))
        ws.cell(row=3, column=i).border = Border(left=Side(border_style='thick', color='FF000000'),
                                                 right=Side(border_style='thick', color='FF000000'),
                                                 top=Side(border_style='thick', color='FF000000'),
                                                 bottom=Side(style='thick'))
    for i in range(3, 5):
        ws['B3'] = 'НЕПРОЦЕНТНІ ДОХОДИ'
        ws['B3'].font = Font(name='Times New Roman', size=16, bold=True)
        ws['B3'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B' + str(i)].border = Border(left=Side(border_style='thick', color='FF000000'),
                                                  right=Side(border_style='thick', color='FF000000'),
                                                  top=Side(border_style='thick', color='FF000000'),
                                                  bottom=Side(style='thin'))

    rd = ws.row_dimensions[3]
    rs = ws.row_dimensions[4]
    rd.height = 45
    rs.height = 30

    for i in range(3, 60):
        ws.cell(row=4, column=i).alignment = Alignment(horizontal='center', vertical='center')


    def slice(x, y, z, limit, lst):
        while y != limit:
            ws.merge_cells(start_row=3, start_column=x, end_row=3, end_column=y)
            x += 3
            y += 3
        for i in range(len(lst)):
            ws.cell(row=3, column=z).value = lst[i]
            ws.cell(row=3, column=z).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row=3, column=z).font = Font(name='Times New Roman', size=12, bold=True)
            ws.cell(row=3, column=z).border = Border(left=Side(border_style='thick', color='FF000000'),
                                                     right=Side(border_style='thick', color='FF000000'),
                                                     top=Side(border_style='thick', color='FF000000'),
                                                     bottom=Side(style='thick'))
            ws.cell(row=3, column=60).border = Border(left=Side(border_style='thick', color='FF000000'),
                                                     right=Side(border_style='thick', color='FF000000'),
                                                     top=Side(border_style='thick', color='FF000000'),
                                                     bottom=Side(style='thick'))
            ws.cell(row=3, column=17).border = Border(left=Side(border_style='thick', color='FF000000'),
                                                      right=Side(border_style='thick', color='FF000000'),
                                                      top=Side(border_style='thick', color='FF000000'),
                                                      bottom=Side(style='thick'))
            z += 3

    slice(3, 5, 3, 20, list_slice)
    slice(19, 21, 19, 63, lst_nazva[:-1])


    freez_obl = ws['C5']
    ws.freeze_panes = freez_obl



    wb.save(file_NII)


formating('NII - test year.xlsx', f'NII на {correct_date} накопичувальний.xlsx')
formating('NII - test month.xlsx', f'NII на {correct_date}.xlsx')


# удаление заготовочных файлов
path = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'NII - test year.xlsx.')
path2 = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'NII - test month.xlsx.')
os.remove(path)
os.remove(path2)
# Считывание файлов за текущий месяц'

wb_1 = openpyxl.load_workbook(f'NII на {correct_date}.xlsx')
wb_2 = openpyxl.load_workbook(f'NII на {correct_date} накопичувальний.xlsx')
wsheet_2 = wb_2['НЕПРОЦЕНТНІ ДОХОДИ']
wsheet_2._parent = wb_1
wb_1._add_sheet(wsheet_2)
wsheet_2.title = '№ накопич'
wb_1.save('Slice.xlsx')


print('Разделение отчета NII на департаменты \n')

def slice_three_dep(num, f_n, s_n, t_n):
    w_b = openpyxl.load_workbook('Slice.xlsx')
    ws_1 = w_b['НЕПРОЦЕНТНІ ДОХОДИ']
    ws_2 = w_b['№ накопич']
    ws_1.title = num
    ws_2.title = num + str(' накопичувальний')


    ws_1['S2'], ws_1['F2'], ws_1['F2'], ws_1['F2'], = '', '', '', ''


    ws_1.delete_cols(3, f_n)
    ws_1.delete_cols(6, s_n)
    ws_1.delete_cols(t_n, 81)

    ws_2.delete_cols(3, f_n)
    ws_2.delete_cols(6, s_n)
    ws_2.delete_cols(t_n, 81)

    w_b.save(f'{num} {correct_date} NII.xlsx')


slice_three_dep('1.38.', 3, 10, 12)
slice_three_dep('1.50.', 6, 13, 9)
slice_three_dep('1.20.', 9, 13, 9)



def slice_other(num, f_n):
    w_b = openpyxl.load_workbook('Slice.xlsx')
    ws_1 = w_b['НЕПРОЦЕНТНІ ДОХОДИ']
    ws_2 = w_b['№ накопич']
    ws_1.title = num
    ws_2.title = num + str(' накопичувальний')

    ws_1.delete_cols(3, f_n)
    ws_1.delete_cols(6, 81)

    ws_2.delete_cols(3, f_n)
    ws_2.delete_cols(6, 81)

    w_b.save(f'{num} {correct_date} NII.xlsx')


slice_other('0.48.', 43)
slice_other('1.06.', 31)
slice_other('1.10.', 34)
slice_other('1.29.', 40)
slice_other('1.46.', 52)
slice_other('1.49.', 49)
slice_other('1.51.', 28)
slice_other('1.55.', 37)
slice_other('1.56.', 46)

path_2 = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'Slice.xlsx')
os.remove(path_2)


total_wb = openpyxl.load_workbook(f'NII на {correct_date}.xlsx')
total_wb_2 = openpyxl.load_workbook(f'NII на {correct_date} накопичувальний.xlsx')
total_wsheet_2 = total_wb_2['НЕПРОЦЕНТНІ ДОХОДИ']
total_wsheet_2._parent = total_wb
total_wb._add_sheet(total_wsheet_2)
total_wsheet_2.title = 'НЕПРОЦЕНТНІ ДОХОДИ накопич'
total_wb.save(f'{correct_date} NII Total.xlsx')


print('Процесс успешно завершен \n')