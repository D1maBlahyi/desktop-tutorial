from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import pandas as pd
from allocation_func import all_func
from datetime import datetime

startTime = datetime.now()
pd.options.mode.chained_assignment = None

pd.set_option('display.max_rows', 5000)
pd.set_option('display.max_columns', 5000)
pd.set_option('display.width', 5000)


def person(name_bi_, name_alloc_, sheet):

    shablon = pd.read_excel('Шаблон.xlsx')
    business = pd.read_excel(name_bi_)

    # датафрейм с планом 5.3. и 5.4.
    state_alloc = business[business['Наименование бизнеса'] == 'Банк'].reset_index(drop=True)

    business = business.loc[~business['Наименование бизнеса'].isin(['Банк'])].reset_index(drop=True)
    business['план'] = business['план'].fillna(0)
    business['факт'] = business['факт'].fillna(0)
    business['виконання %'] = 0
    for i in range(len(business)):
        if business['план'][i] != 0:
            business['виконання %'][i] = round(business['факт'][i] / business['план'][i] * 100)
        else:
            business['виконання %'][i] = 100
    business['план'] = business['план'].replace(1, 0)
    business['відхилення'] = business['факт'] - business['план']

    table_names = business['Наименование бизнеса'].drop_duplicates()
    table_names = list(sorted(table_names))


    def transposition(value):
        wide_titanic = business.pivot(
            index=['Код статьи', 'Наименование статьи'],
            columns='Наименование бизнеса',
            values=value
        )
        wide_titanic.reset_index(drop=False, inplace=True)
        return wide_titanic

    lst_num = ["план", 'факт', "відхилення", "виконання %"]

    concat_df = pd.concat([transposition(lst_num[i]) for i in range(len(lst_num))], axis=1, ignore_index=True)
    concat_df = concat_df.rename(columns={0: 'newName0', 1: 'newName1'})

    def moving(x):
        data = concat_df.iloc[:, [x, x + 7, x + 14, x + 21]]
        return data

    lst_count_buss = [2, 3, 4, 5, 6]


    concat_df_ = pd.concat([moving(lst_count_buss[i]) for i in range(len(lst_count_buss))], axis=1, ignore_index=True)
    concat_df_2 = pd.concat([concat_df.iloc[:, :2], concat_df_], axis=1)

    alone_lst = list(
        [concat_df_2.iloc[i, 0] for i in range(len(concat_df_2)) if concat_df_2.iloc[i, 0].startswith('4.11')
         or concat_df_2.iloc[i, 0].startswith('5.11')]) + ['4.10.10.', '5.10.10.']

    # важные (датафрейм с трансфертами и без) ====================
    alone_df = concat_df_2.loc[concat_df_2.iloc[:, 0].isin(alone_lst)]
    concat_df_2 = concat_df_2.loc[~concat_df_2.iloc[:, 0].isin(alone_lst)].reset_index(drop=True)
    # ==================================

    # Статьи которые в шаблоне и в базе, но как цветные ====================
    lst_shablon = []
    for i in range(len(concat_df_2)):
        if concat_df_2.iloc[i, 0] in shablon['Код статьи'].to_list():
            lst_shablon.append(concat_df_2.iloc[i, 0])
    shablon_for_merge = concat_df_2.loc[concat_df_2.iloc[:, 0].isin(lst_shablon)]


    # датафрейм базы очищенный
    concat_df_2 = concat_df_2.loc[~concat_df_2.iloc[:, 0].isin(lst_shablon)].reset_index(drop=True)

    CHek = shablon['Код статьи'][shablon['Чек'] != 1.0].to_list()
    chek_2 = shablon['Код статьи'][shablon['Чек_2'] == 1.0].to_list()
    chek_3 = shablon['Код статьи'][shablon['Чек_3'] == 1.0].to_list()

    df_lst_map = []
    for i in range(len(CHek)):
        second_df = []
        if CHek[i] not in chek_2:
            df_lst_map.append(CHek[i])
        elif CHek[i] in chek_2:
            for j in range(len(shablon)):
                if shablon['Код статьи'][j].startswith(CHek[i]):
                    if shablon['Код статьи'][j] not in chek_3:
                        second_df.append(shablon['Код статьи'][j])
            for k in range(len(concat_df_2)):
                if concat_df_2['newName0'][k].startswith(CHek[i]):
                    second_df.append(concat_df_2['newName0'][k])
            second_df = sorted(second_df)
            for h in range(len(second_df)):
                df_lst_map.append(second_df[h])

    # Полный порядок только статтей
    data = pd.DataFrame(data={'Код статьи': df_lst_map})
    data = data.drop_duplicates(keep='first').reset_index(drop=True)

    def nlo(state, data, chek_2):
        lst_key = []
        lst_val = []
        for i in range(len(data)):
            if data['Код статьи'][i].startswith(state):
                lst_key.append(i)
                lst_val.append(data['Код статьи'][i])
        if len(lst_key) > 2:
            while len(lst_key) > 2:
                data = data.drop(lst_key[-1], axis=0)
                line = pd.DataFrame({"Код статьи": lst_val[-1]}, index=[lst_key[1] + 0.5])
                data = data.append(line, ignore_index=False)
                data = data.sort_index().reset_index(drop=True)
                chek_2.append(lst_val[-1])
                lst_key.pop(-1)
                lst_val.pop(-1)
        return data

    data = nlo('5.3.6.38.', data, chek_2)
    data = nlo('5.4.6.5.26.', data, chek_2)

    lst_rez = []
    lst_rez_drop = []

    for i in range(len(data)):
        if data['Код статьи'][i] == '5.5.7.2.1.2.' or data['Код статьи'][i] == '5.5.7.2.2.2.':
            data = data.drop(i, axis=0)
    data = data.reset_index(drop=True)

    for i in range(len(data)):
        for j in range(4, 8):
            if data['Код статьи'][i] == f'5.5.7.{j}.':
                lst_rez.append(data['Код статьи'][i])
                lst_rez_drop.append(i)


    for i in range(len(lst_rez_drop)):
        data = data.drop(lst_rez_drop[i], axis=0
                         )
    data = data.reset_index(drop=True)

    line = 0
    for i in range(len(data)):
        if data['Код статьи'][i] == '5.5.7.':
            line = pd.DataFrame({"Код статьи": '5.5.7.2.1.2.'}, index=[i + 0.5])
            data = data.append(line, ignore_index=False)
            data = data.sort_index().reset_index(drop=True)

    for i in range(len(data)):
        if data['Код статьи'][i] == '5.5.7.2.1.2.':
            line = pd.DataFrame({"Код статьи": '5.5.7.2.2.2.'}, index=[i + 0.5])
            data = data.append(line, ignore_index=False)
            data = data.sort_index().reset_index(drop=True)


    lst_rez = lst_rez[::-1]

    for i in range(len(data)):
        if data['Код статьи'][i] == "#16":
            for j in range(len(lst_rez)):
                line = pd.DataFrame({"Код статьи": lst_rez[j]}, index=[i + 0.5])
                data = data.append(line, ignore_index=False)
                data = data.sort_index().reset_index(drop=True)


    data = data.drop_duplicates().reset_index(drop=True)


    # Датафрейм с наименования шаблона (без базы)
    merge_tab = pd.merge(left=data, right=shablon, left_on='Код статьи', right_on='Код статьи', how='left')
    merge_tab = merge_tab[['Код статьи', 'Наименование']]

    # Обьедененная таблица с цветными и базой
    df_values = pd.concat([concat_df_2, shablon_for_merge], axis=0)

    # Полный комплект
    all_merge = pd.merge(left=merge_tab, right=df_values, left_on='Код статьи', right_on='newName0', how='left')
    all_merge['Наименование'] = all_merge['Наименование'].fillna(all_merge['newName1'])
    all_merge = all_merge.drop(['newName0', 'newName1'], axis=1)






    def func(tini_df):
        tini_df['план_'] = tini_df.iloc[:, 2::4].sum(axis=1)
        tini_df['факт_'] = tini_df.iloc[:, 3::4].sum(axis=1)
        tini_df['відхилення_'] = tini_df.iloc[:, 4::4].sum(axis=1)
        tini_df['виконання %_'] = tini_df.iloc[:, 5::4].sum(axis=1)

        tini_df['_план_'] = tini_df.iloc[:, [2, 10]].sum(axis=1)
        tini_df['_факт_'] = tini_df.iloc[:, [3, 11]].sum(axis=1)
        tini_df['_відхилення_'] = tini_df.iloc[:, [4, 12]].sum(axis=1)
        tini_df['_виконання %_'] = tini_df.iloc[:, [5, 13]].sum(axis=1)

        consol_bank = tini_df.iloc[:, 22:26]
        corp_buss = tini_df.iloc[:, 26:30]
        corp_net = tini_df.iloc[:, 10:14]
        vip = tini_df.iloc[:, 2:6]
        msb = tini_df.iloc[:, 14:18]
        rozdrib = tini_df.iloc[:, 18:22]
        inshi = tini_df.iloc[:, 6:10]

        conclusion = pd.concat([tini_df.iloc[:, [0, 1]], consol_bank, corp_buss, corp_net, vip, msb, rozdrib, inshi],
                               axis=1)
        return conclusion

    conclusion = func(all_merge)

    lst_columns = ['Код статьи', 'Наименование статьи'] + lst_num * (len(table_names) + 2)
    conclusion.columns = lst_columns



    # Расчет трансфертов
    transfert = func(alone_df)
    transfert.columns = lst_columns
    transfert = transfert.reset_index(drop=True)

    transfert_dohid_4 = ('4.11.', '4.10.10.')
    first_trans = transfert[transfert['Код статьи'].str.startswith(transfert_dohid_4)]
    first_trans.loc['Row_plan'] = {'Код статьи': '$', 'Наименование статьи': 'Трансферти_4.11.'}
    first_trans.iloc[-1, 2:] = first_trans.iloc[:, 2:].sum(axis=0)
    first_trans.iloc[-1, 2:5] = 0

    transfert_vytrat_5 = ('5.11.', '5.10.10.')
    second_trans = transfert[transfert['Код статьи'].str.startswith(transfert_vytrat_5)]
    second_trans.loc['Row_plan'] = {'Код статьи': '$', 'Наименование статьи': 'Трансферти_5.11.'}
    second_trans.iloc[-1, 2:] = second_trans.iloc[:, 2:].sum(axis=0)
    second_trans.iloc[-1, 2:5] = 0



    conclusion = conclusion.fillna(0)
    for i in range(len(conclusion)):
        if conclusion['Код статьи'][i] == '!1':
            conclusion.iloc[i, 2:] += second_trans.iloc[-1, 2:]
        elif conclusion['Код статьи'][i] == '!2':
            conclusion.iloc[i, 2:] += first_trans.iloc[-1, 2:]


    for i in range(len(conclusion)):
        for j in range(len(state_alloc)):
            if conclusion['Код статьи'][i] == state_alloc['Код статьи'][j]:
                conclusion.iloc[i, 2] = state_alloc.iloc[j, 6]

    allocation = pd.read_excel(name_alloc_)
    allocation = allocation.loc[~allocation['Функционал-юнит (источник)'].isin(['нет в'])].reset_index(
        drop=True)
    all_for_bus = ('1.6.', '1.4.')
    allocation = allocation[allocation['Тип связи'].str.startswith(all_for_bus)].reset_index(drop=True)
    allocation = allocation.iloc[:, [0, 4, 6, 7]]
    table_names_all = allocation['Наименование бизнеса'].drop_duplicates()
    table_names_all = list(sorted(table_names_all))
    name_all = 0
    for i in range(len(table_names_all)):
        if len(table_names_all) < len(table_names):
            name_all = "".join(list(set(table_names) - set(table_names_all)))
    allocation.loc[-1] = {'Тип связи': '1.4.', 'Наименование бизнеса': name_all, 'факт': 0, 'план': 0}
    allocation.reset_index(drop=True, inplace=True)
    allocation.loc[-1] = {'Тип связи': '1.6.', 'Наименование бизнеса': name_all, 'факт': 0, 'план': 0}
    allocation.reset_index(drop=True, inplace=True)

    allocation['план'] = allocation['план'].fillna(0)
    allocation['факт'] = allocation['факт'].fillna(0)

    aggr_data = allocation.groupby(by=['Тип связи', 'Наименование бизнеса'])[
        ['факт', 'план']].sum()
    aggr_data = aggr_data.reset_index(drop=False)


    aggr_data['відхилення'] = aggr_data['факт'] - aggr_data['план']
    aggr_data['виконання %'] = 0
    for i in range(len(aggr_data)):
        if aggr_data['план'][i] != 0:
            aggr_data['виконання %'][i] = round(aggr_data['факт'][i] / aggr_data['план'][i] * 100)
        else:
            aggr_data['виконання %'][i] = 100





    def transposition_2(value):
        wide_titanic = aggr_data.pivot(
            index=['Тип связи'],
            columns='Наименование бизнеса',
            values=value
        )
        wide_titanic.reset_index(drop=False, inplace=True)
        return wide_titanic

    all_pivot = pd.concat([transposition_2(lst_num[i]) for i in range(len(lst_num))], axis=1, ignore_index=True)
    all_pivot = all_pivot.rename(columns={0: 'newName0'})


    def moving_2(x):
        data = all_pivot.iloc[:, [x, x + 6, x + 12, x + 18]]
        return data

    lst_count_buss = [1, 2, 3, 4, 5]

    refresh_all = pd.concat([moving_2(lst_count_buss[i]) for i in range(len(lst_count_buss))], axis=1,
                            ignore_index=True)

    refresh_all = pd.concat([all_pivot.iloc[:, :1], all_pivot.iloc[:, :1], refresh_all], axis=1)


    stroki_all = func(refresh_all)
    lst_columns_all = ['Код статьи', 'Наименование статьи'] + lst_num * (len(table_names) + 2)
    stroki_all.columns = lst_columns_all
    stroki_all.iloc[:, 2:6] = 0


    for i in range(len(conclusion)):
        for j in range(len(stroki_all)):
            if conclusion['Код статьи'][i] == stroki_all['Код статьи'][j]:
                conclusion.iloc[i, 2:] = stroki_all.iloc[j, 2:]
    def counter(prep_num, next_num):
        count = next_num
        zone_1 = prep_num
        try:
            while len(conclusion['Код статьи'][prep_num].split('.')) < \
                    len(conclusion['Код статьи'][next_num].split('.')) and conclusion['Код статьи'][prep_num].split(
                '.')[:2] == \
                    conclusion['Код статьи'][next_num].split('.')[:2]:
                count += 1
                next_num += 1
        except:
            print('-')

        conclusion.iloc[prep_num, 2:] = conclusion.iloc[zone_1:count, 2:].sum(axis=0)

        return conclusion

    tini_df = 0
    for i in range(len(conclusion) - 2):
        tini_df = counter(i, i + 1)



    name_code = tini_df[['Наименование статьи']]
    name_state = tini_df[['Код статьи']]
    tini_df = tini_df.drop(['Наименование статьи'], axis=1)
    tini_df.set_index('Код статьи', inplace=True)

    tini_df.loc['#1'] = tini_df.loc['4.1.'] + tini_df.loc['!1']
    tini_df.loc['#2'] = tini_df.loc['5.1.'] + tini_df.loc['!2']
    tini_df.loc['!3'] = tini_df.loc['5.5.1.'] + tini_df.loc['5.5.3.']
    tini_df.loc['#3'] = tini_df.loc['#1'] + tini_df.loc['#2']
    tini_df.loc['#4'] = tini_df.loc['#3'] + tini_df.loc['!3']
    tini_df.loc['#5'] = tini_df.loc['4.2.'] + tini_df.loc['5.2.']
    tini_df.loc['#6'] = tini_df.loc['4.3.'] + tini_df.loc['5.5.4.4.']
    tini_df.loc['#7'] = tini_df.loc['4.10.'] + tini_df.loc['5.10.']
    tini_df.loc['5.5.7.'] = tini_df.loc['5.5.7.2.1.2.'] + tini_df.loc['5.5.7.2.2.2.']
    tini_df.loc['#8'] = tini_df.loc['4.8.'] + tini_df.loc['4.9.'] + tini_df.loc['4.6.'] + tini_df.loc['4.7.']

    if '4.3.1.2.' in tini_df.index:
        print('yes')
        tini_df.loc['#9'] = tini_df.loc['#5'] + tini_df.loc['4.3.'] - (tini_df.loc['4.3.1.2.'] + tini_df.loc['4.3.2.2.'] +
                                                                   tini_df.loc['4.3.3.2.'] + tini_df.loc['4.3.4.2.'] +
                                                                   tini_df.loc['4.3.5.2.'] + tini_df.loc['4.3.6.2.']) + \
                        tini_df.loc['#7'] + tini_df.loc['4.4.'] + tini_df.loc['4.5.'] + tini_df.loc['#8'] - \
                        tini_df.loc['4.9.17.'] - tini_df.loc['4.5.2.'] - tini_df.loc['5.3.10.'] - tini_df.loc['4.8.2.1.1.2.']
    else:
        tini_df.loc['#9'] = tini_df.loc['#5'] + tini_df.loc['4.3.'] - (
                    tini_df.loc['4.3.2.2.'] +
                    tini_df.loc['4.3.3.2.'] + tini_df.loc['4.3.4.2.'] +
                    tini_df.loc['4.3.5.2.'] + tini_df.loc['4.3.6.2.']) + \
                            tini_df.loc['#7'] + tini_df.loc['4.4.'] + tini_df.loc['4.5.'] + tini_df.loc['#8'] - \
                            tini_df.loc['4.9.17.'] - tini_df.loc['4.5.2.'] - tini_df.loc['5.3.10.'] - tini_df.loc[
                                '4.8.2.1.1.2.']
    tini_df.loc['!4'] = tini_df.loc['5.5.7.1.'] + tini_df.loc['5.5.7.2.'] + tini_df.loc['5.5.7.']
    tini_df.loc['#10'] = tini_df.loc['4.13.'] + tini_df.loc['4.8.11.'] + tini_df.loc['5.13.'] + tini_df.loc['5.3.13.']
    tini_df.loc['#11'] = tini_df.loc['4.8.9.25.'] + tini_df.loc['4.9.16.'] + tini_df.loc['5.3.6.38.'] + tini_df.loc[
        '5.4.6.5.26.']
    tini_df.loc['#12'] = tini_df.loc['#3'] + tini_df.loc['#5'] + tini_df.loc['#6'] + tini_df.loc['#7'] + \
                         tini_df.loc['4.4.'] + tini_df.loc['4.14.'] + \
                         tini_df.loc['4.5.'] + tini_df.loc['#8'] + tini_df.loc['#10'] + tini_df.loc['#11']
    tini_df.loc['#13'] = tini_df.loc['#12'] + tini_df.loc['!4'] + tini_df.loc['5.5.4.4.'] + tini_df.loc['!3']
    tini_df.loc['!5'] = tini_df.loc['4.10.4.'] + tini_df.loc['4.10.5.'] + tini_df.loc['4.10.6.']
    tini_df.loc['!6'] = tini_df.loc['5.10.4.'] + tini_df.loc['5.10.5.'] + tini_df.loc['5.10.6.']
    tini_df.loc['#14'] = tini_df.loc['!5'] + tini_df.loc['!6']
    tini_df.loc['#15'] = tini_df.loc['5.3.'] + tini_df.loc['5.4.'] + tini_df.loc['1.6.'] + tini_df.loc['1.4.']
    if len(lst_rez) > 1:
        tini_df.loc['#16'] = tini_df.loc['5.5.8.'] + tini_df.loc['5.5.7.4.'] + tini_df.loc['5.5.7.6.']
    else:
        tini_df.loc['#16'] = tini_df.loc['5.5.8.'] + tini_df.loc['5.5.7.4.']
    tini_df.loc['#17'] = tini_df.loc['#16'] + tini_df.loc['#15'] + tini_df.loc['#14'] + tini_df.loc['#13']

    tini_df.reset_index(
        drop=False,  # False означает, что существующий индекс переместится в колонки, а не будет удален
        inplace=True  # произвести операцию на месте, без присваивания новой переменной
    )

    res_finish_ = pd.concat([name_state, name_code, tini_df.iloc[:, 1:]], axis=1)
    # if __name__ == '__main__':

    all_finish = all_func(name_alloc_)
    res_finish_ = res_finish_

    save_col = res_finish_.columns
    res_finish_.columns = list(range(1, 31))
    all_finish.columns = list(range(1, 31))
    res = pd.concat([res_finish_, all_finish], axis=0).reset_index(drop=True)
    res.columns = save_col
    res.set_index('Код статьи', inplace=True)
    res.loc['#18'] = res.loc['#17'] + res.loc['!7'] + res.loc['5.6.']
    res.iloc[len(res) - 1, 0] = 'ПРИБУТОК після аллокацій та оподаткування'
    res.reset_index(
        drop=False,  # False означает, что существующий индекс переместится в колонки, а не будет удален
        inplace=True  # произвести операцию на месте, без присваивания новой переменной
    )

    num_col = 2
    while num_col <= 27:
        for i in range(len(res)):
            res.iloc[i, num_col + 2] = res.iloc[i, num_col + 1] - res.iloc[i, num_col]
            if -1 < res.iloc[i, num_col] < 1:
                res.iloc[i, num_col] = 0
            if res.iloc[i, num_col] == 0:
                if res.iloc[i, num_col + 1] == 0:
                    res.iloc[i, num_col + 3] = '-'
                elif res.iloc[i, num_col + 1] > 0:
                    res.iloc[i, num_col + 3] = '300%+'
                elif res.iloc[i, num_col + 1] < 0:
                    res.iloc[i, num_col + 3] = '- 300%+'
            elif res.iloc[i, num_col + 1] == 0:
                if res.iloc[i, num_col] > 0:
                    res.iloc[i, num_col + 3] = '0%'
                elif res.iloc[i, num_col] < 0:
                    res.iloc[i, num_col + 3] = '300%+'
            elif res.iloc[i, num_col + 1] < 0 and res.iloc[i, num_col] < 0:
                res.iloc[i, num_col + 3] = res.iloc[i, num_col] / res.iloc[i, num_col + 1]
                if res.iloc[i, num_col + 3] > 3:
                    res.iloc[i, num_col + 3] = '300%+'
            elif res.iloc[i, num_col + 1] > 0 and res.iloc[i, num_col] > 0:
                res.iloc[i, num_col + 3] = res.iloc[i, num_col + 1] / res.iloc[i, num_col]
                if res.iloc[i, num_col + 3] > 3:
                    res.iloc[i, num_col + 3] = '300%+'
            else:
                res.iloc[i, num_col + 3] = '300%+'
        num_col += 4


    wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(res, index=False):
        ws.append(r)

    for column in ws.columns:
        lenght = max(len(str(cell.value)) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = lenght

    union = chek_3 + ['#18']
    blue = []
    grey = chek_2
    thin = []
    all_color = []
    for i in range(len(all_finish)):
        if len(all_finish.iloc[i, 0]) > 4:
            all_color.append(all_finish.iloc[i, 0])

    for i in range(len(union)):
        if '#' in union[i]:
            blue.append(union[i])
        elif '4.4.' in union[i]:
            blue.append(union[i])
        elif '4.14.' in union[i]:
            blue.append(union[i])
        elif '4.5.' in union[i]:
            blue.append(union[i])
    for word in blue:
        union.remove(word)

    for i in range(len(res)):
        if res['Код статьи'][i] not in union and res['Код статьи'][i] not in grey and res['Код статьи'][i] not in blue:
            thin.append(res['Код статьи'][i])

    all_col = ['1.2.', '1.7.', '1.1.']
    thin.remove('!7')
    thin.remove('1.2.')
    thin.remove('1.7.')
    thin.remove('1.1.')


    lists = [[] for _ in range(12)]
    list_all_gr = []
    list_alloc_2 = []

    col_range = ws.max_column

    for i in range(1, len(res.values) + 2):

        for col in range(1, col_range + 1):
            cell_header = ws.cell(i, col)
            cell_header.alignment = Alignment(horizontal='right', vertical='center')
            cell_header.font = Font(name='Times New Roman', size=10, bold=True)
            cell_header.border = Border(left=Side(border_style='thin', color='FF000000'),
                                        right=Side(border_style='thin', color='FF000000'),
                                        top=Side(border_style='thin', color='FF000000'),
                                        bottom=Side(style='thin'))
        ws['B' + str(i)].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws['A' + str(i)].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)


        names = [ws['B' + str(i)].value]
        indexes = [ws['A' + str(i)].value]

        def main_format(color, size):
            for col in range(1, col_range + 1):
                cell_header = ws.cell(i, col)
                cell_header.fill = PatternFill(start_color=color, fill_type="solid")
                cell_header.font = Font(name='Times New Roman', size=size, bold=True)
                cell_header.number_format = '#,##0'
            ws['B' + str(i)] = ws['B' + str(i)].value.upper()
            ws['B' + str(i)].font = Font(name='Times New Roman', size=size, bold=True)
            ws['A' + str(i)].font = Font(name='Times New Roman', size=size, italic=True, bold=True)
            ws['B' + str(i)].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws['A' + str(i)].fill = PatternFill(fill_type='solid', start_color=color)
            ws['B' + str(i)].fill = PatternFill(fill_type='solid', start_color=color)
            for col in range(6, col_range + 1, 4):
                cell_header = ws.cell(i, col)
                cell_header.number_format = '0%'

        def low_format(color):
            for col in range(1, col_range + 1):
                cell_header = ws.cell(i, col)
                cell_header.fill = PatternFill(start_color=color, fill_type="solid")
                cell_header.font = Font(name='Times New Roman', size=9, italic=True, bold=False)
                cell_header.number_format = '#,##0'
            ws['B' + str(i)].font = Font(name='Times New Roman', size=9, bold=False)
            ws['B' + str(i)].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws['A' + str(i)].font = Font(name='Times New Roman', size=9, bold=False)
            ws['A' + str(i)].fill = PatternFill(fill_type='solid', start_color=color)
            ws['B' + str(i)].fill = PatternFill(fill_type='solid', start_color=color)
            for col in range(6, col_range + 1, 4):
                cell_header = ws.cell(i, col)
                cell_header.number_format = '0%'

        for j in indexes:
            if j in blue:
                main_format('B0E0E6', 12)
                lists[0].append(i)
            elif j in union:
                main_format('3CB371', 11)
                lists[1].append(i)
            elif j in grey:
                main_format('C0C0C0', 10)
                lists[2].append(i)
            elif j == '!7':
                main_format('FFB233', 12)
            elif len(j.split('.')[:-1]) == 4:
                low_format('BD5E86')
                lists[3].append(i)
            elif len(j.split('.')[:-1]) == 5:
                low_format('DDE533')
                lists[4].append(i)
            elif len(j.split('.')[:-1]) == 6:
                low_format('90A8B8')
                lists[5].append(i)
            elif len(j.split('.')[:-1]) == 7:
                low_format('D1C1E2')
                lists[6].append(i)
            elif len(j.split('.')[:-1]) == 8:
                low_format('E1B0A2')
                lists[7].append(i)
            elif len(j.split('.')[:-1]) == 9:
                low_format('A44953')
                lists[8].append(i)
            elif len(j.split('.')[:-1]) == 10:
                low_format('5D8668')
                lists[9].append(i)
            elif j in all_col:
                main_format('E5B9EB', 11)
                list_all_gr.append(i)
            elif j in all_color:
                low_format('FFFFFF')
                list_alloc_2.append(i)

    for i in range(len(lists[0])):
        lists[0][i] += 2
    for i in range(0, len(lists[0]) - 1, 1):
        ws.row_dimensions.group(lists[0][i] + 1, lists[0][i + 1] - 1, hidden=True, outline_level=1)
    for i in range(len(list_all_gr)):
        list_all_gr[i] += 2
    for i in range(0, len(list_all_gr) - 1, 1):
        ws.row_dimensions.group(list_all_gr[i] + 1, list_all_gr[i + 1] - 1, hidden=True, outline_level=3)
    ws.row_dimensions.group(list_all_gr[-1] + 1, lists[0][-1] - 1, hidden=True, outline_level=3)

    def dimension(lst, lvl, step, lst_2):
        for i in range(len(lst)):
            lst[i] += 2
        for i in range(0, len(lst)-1, step):
            for j in range(1, len(lst_2)-1):
                if lst[i] < lst[i+1] < lst_2[j]:
                    ws.row_dimensions.group(lst[i]+1, lst[i+1]-1, hidden=True, outline_level=lvl)
                    break
                elif lst[i] < lst_2[j] < lst[i + 1]:
                    ws.row_dimensions.group(lst[i] + 1, lst_2[j] - 1, hidden=True, outline_level=lvl)
                    break

    dimension(lists[1], 2, 1, lists[0])
    lists[1] = list(sorted(lists[0] + lists[1]))

    dimension(lists[2], 3, 1, lists[1])
    lists[2] = list(sorted(lists[1] + lists[2]))

    dimension(lists[3], 4, 1, lists[2])
    lists[3] = list(sorted(lists[2] + lists[3]))

    dimension(lists[4], 5, 1, lists[3])



    for i in range(1, len(res.values) + 2):
        indexes = [ws['A' + str(i)].value]
        for j in indexes:
            if '#' in j or '!' in j:
                ws['A' + str(i)] = np.nan

    ws.insert_rows(0)
    ws.insert_rows(0)

    for i in range(3, 30):
        ws.cell(row=3, column=i).alignment = Alignment(horizontal='center', vertical='center')

    lst_names = ['Консолідовано Банк', 'Корпоративний бізнес', 'Корпоратив - мережа', 'VIP Корпоратив',
                 'Малий та середній бізнес', 'Роздрібний  банкінг', 'Інші']

    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)

    for i in range(2, 4):
        ws['A2'] = 'Код статей'
        ws['A2'].font = Font(name='Times New Roman', size=14, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A' + str(i)].border = Border(left=Side(border_style='thick', color='FF000000'),
                                                  right=Side(border_style='thick', color='FF000000'),
                                                  top=Side(border_style='thick', color='FF000000'),
                                                  bottom=Side(style='thin', color='FF000000'))
        ws['B2'] = 'Найменування статей'
        ws['B2'].font = Font(name='Times New Roman', size=14, bold=True)
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B' + str(i)].border = Border(left=Side(border_style='thick', color='FF000000'),
                                         right=Side(border_style='thick', color='FF000000'),
                                         top=Side(border_style='thick', color='FF000000'),
                                         bottom=Side(style='thin', color='FF000000'))
        rd = ws.row_dimensions[2]
        rs = ws.row_dimensions[3]
        rd.height = 35
        rs.height = 25
        rq = ws.row_dimensions[1]
        rq.height = 25

    for i in range(3, 29):
        ws.cell(row=2, column=i).border = Border(left=Side(border_style='thick', color='FF000000'),
                                                 right=Side(border_style='thick', color='FF000000'),
                                                 top=Side(border_style='thick', color='FF000000'),
                                                 bottom=Side(style='thick', color='FF000000'))

    def slice(x, y, z, limit, lst):
        while y <= limit:
            ws.merge_cells(start_row=2, start_column=x, end_row=2, end_column=y)
            x += 4
            y += 4
        for i in range(len(lst)):
            ws.cell(row=2, column=z).value = lst[i]
            ws.cell(row=2, column=z).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row=2, column=z).font = Font(name='Times New Roman', size=12, bold=True)

            z += 4


    slice(3, 6, 3, len(res.columns), lst_names)

    for column in ws.columns:
        lenght = max(len(str(cell.value)) for cell in column)
        if lenght > 75:
            lenght = 100
        ws.column_dimensions[column[0].column_letter].width = lenght

    ws['A1'] = sheet + '.2023'

    freez_obl = ws['C4']
    ws.freeze_panes = freez_obl
    ws.title = sheet

    wb.save(f"C:\\Users\\admin\PycharmProjects\\fin_b\\files\\{sheet}_2023.xlsx")
    wb.close()
    print(f"Час вигрузки задачі {sheet}", datetime.now() - startTime)
    res.to_excel(f'res_{sheet}.xlsx', index=False)

    return res



