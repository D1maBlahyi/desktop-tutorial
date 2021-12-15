import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import openpyxl




def tables(units, business):

    pd.set_option('display.max_rows', 5000)
    pd.set_option('display.max_columns', 5000)
    pd.set_option('display.width', 5000)

    # Заголовочная часть таблицы
    ostatok_1_units = units[['№ (рівень 1)', 'Назва (рівень 1)', '№ (рівень 2)', 'Назва (рівень 2)', '№ (рівень 3)',
                             'Назва (рівень 3)', 'Група статей (рівень.4)',
         'Наименование группы 4', 'БАНК (п.), грн.', 'БАНК (ф.), грн.', 'БАНК (%)']]

    # Значения с таблицы бизнесов
    business.drop(
        ['№ (рівень 1)', 'Назва (рівень 1)', '№ (рівень 2)', 'Назва (рівень 2)', '№ (рівень 3)', 'Назва (рівень 3)',
         'Група статей (рівень.4)',
         'Наименование группы 4', 'БАНК (п.), грн.', 'БАНК (ф.), грн.', 'БАНК (%)'],
        axis=1,  # удаление строк происходит аналогично, чтобы удалить именно колонки, выбираем соответствующую ось
        inplace=True  # удаление "на месте", без присваивания новой переменной
    )

    bus_copy = business.copy()

    # Выравнивание значений и замена на пустоту
    for i in range(20, 26):
        bus_copy.loc[i] = np.nan
    nan = bus_copy.drop(bus_copy.index[0:20])
    nan = nan.drop(nan.index[6:len(nan)])
    df = pd.DataFrame(business[20:len(business)])
    df.index = (df.index + 6)
    buss = business.drop(business.index[20:len(business)])
    business1 = pd.concat([buss, nan], axis=0)
    business2 = pd.concat([business1, df], axis=0)


    # Значения с таблицы юнитов
    units.drop(
        ['№ (рівень 1)', 'Назва (рівень 1)', '№ (рівень 2)', 'Назва (рівень 2)', '№ (рівень 3)', 'Назва (рівень 3)',
         'Група статей (рівень.4)',
         'Наименование группы 4', 'БАНК (п.), грн.', 'БАНК (ф.), грн.', 'БАНК (%)'],
        axis=1,  # удаление строк происходит аналогично, чтобы удалить именно колонки, выбираем соответствующую ось
        inplace=True  # удаление "на месте", без присваивания новой переменной
    )

    combiend_df_1 = pd.concat([ostatok_1_units, business2], axis=1)

    # Об'єднана таблиця фул
    combiend_df_2 = pd.concat([combiend_df_1, units], axis=1)
    # ========================================================================================================
    # Таблица для агрегирования
    cd2 = combiend_df_2.copy()
    cd2.drop(
        ['№ (рівень 1)', 'Назва (рівень 1)', '№ (рівень 2)', 'Назва (рівень 2)', '№ (рівень 3)', 'Назва (рівень 3)',
         'Група статей (рівень.4)',
         'Наименование группы 4'],
        axis=1,  # удаление строк происходит аналогично, чтобы удалить именно колонки, выбираем соответствующую ось
        inplace=True  # удаление "на месте", без присваивания новой переменной
    )

    # Агрегирования данных по Комисионным доходам
    combiend_df_2.set_index('№ (рівень 1)', inplace=True)
    riven_4_2 = combiend_df_2.loc['4.2.']

    part1 = riven_4_2.groupby(by=['№ (рівень 1)', 'Назва (рівень 1)'])[cd2.columns].max()
    part2 = riven_4_2.groupby(by=['№ (рівень 2)', 'Назва (рівень 2)'])[cd2.columns].sum()
    part3 = riven_4_2.groupby(by=['№ (рівень 3)', 'Назва (рівень 3)'])[cd2.columns].sum()
    part4 = riven_4_2.groupby(by=['Група статей (рівень.4)', 'Наименование группы 4'])[cd2.columns].sum()

    res = pd.concat([part1, part2, part3, part4], axis=0)

    # Агрегирования данных по Комиссионным затратам
    riven_5_2 = combiend_df_2.loc['5.2.']
    vytraty1 = riven_5_2.groupby(by=['№ (рівень 1)', 'Назва (рівень 1)'])[cd2.columns].min()
    vytraty2 = riven_5_2.groupby(by=['№ (рівень 2)', 'Назва (рівень 2)'])[cd2.columns].sum()
    vytraty3 = riven_5_2.groupby(by=['№ (рівень 3)', 'Назва (рівень 3)'])[cd2.columns].sum()

    res2 = pd.concat([vytraty1, vytraty2, vytraty3], axis=0)

    # Комиссиоонные доходы и затраты
    data = pd.DataFrame(pd.concat([res, res2], axis=0))
    data = data.sort_values(by='№ (рівень 1)')

    # ======================================================================================================================
    # Чистый комиссионный доход
    clear_dohid = data.loc['4.2.', 'Комісійні доходи'] + data.loc['5.2.', 'Комісійні витрати']
    df_cd = pd.DataFrame(clear_dohid).T
    fors = pd.DataFrame(
        data={
            '№ (рівень 1)': [''],
            'Назва (рівень 1)':['Чистий комісійний дохід']
        })
    clear_dohid_full = pd.concat([fors, df_cd], axis=1)


    def zamena_indexa(table):
        table.reset_index(
            drop=False,  # False означает, что существующий индекс переместится в колонки, а не будет удален
            inplace=True  # произвести операцию на месте, без присваивания новой переменной
        )
    zamena_indexa(data)

    data = pd.concat([data, clear_dohid_full])

    # ======================================================================================================================
    # Чистые доходы
    urovni = combiend_df_2.loc['4.3.':'4.5.']
    r1 = urovni.groupby(by=['№ (рівень 1)', 'Назва (рівень 1)'])[cd2.columns].sum()
    zamena_indexa(r1)

    # Другие доходы
    urovni2 = combiend_df_2.loc['4.6.':'4.7.']
    r2 = urovni2.groupby(by=['№ (рівень 1)', 'Назва (рівень 1)'])[cd2.columns].sum()
    zamena_indexa(r2)

    # Другие операционные доходы
    riven_4_8 = combiend_df_2.loc['4.8.']
    insh_op_doh_1 = riven_4_8.groupby(by=['№ (рівень 1)', 'Назва (рівень 1)'])[cd2.columns].max()
    insh_op_doh_2 = riven_4_8.groupby(by=['№ (рівень 2)', 'Назва (рівень 2)'])[cd2.columns].sum()
    res3 = pd.concat([insh_op_doh_1, insh_op_doh_2], axis=0)
    zamena_indexa(res3)


    # Другие доходы
    riven_4_9 = combiend_df_2.loc['4.9.']
    insh_doh_1 = riven_4_9.groupby(by=['№ (рівень 1)', 'Назва (рівень 1)'])[cd2.columns].max()
    insh_doh_2 = riven_4_9.groupby(by=['№ (рівень 2)', 'Назва (рівень 2)'])[cd2.columns].sum()
    res4 = pd.concat([insh_doh_1, insh_doh_2], axis=0)
    zamena_indexa(res4)

    data = pd.concat([data, r1, r2, res3, res4])

    # =======================================================================================================================
    # Внутренние комиссионые доходы и затраты
    vnutrishni = combiend_df_2.loc['4.10.':'5.10.']
    vnutrishni2 = vnutrishni.groupby(by=['№ (рівень 1)', 'Назва (рівень 1)'])[cd2.columns].max()
    vnutrishni2 = vnutrishni2.iloc[:2]

    # Чистые внутренние комиссионые доходы
    clear_vnutri_dohid = vnutrishni2.loc['4.10.', 'Внутрішні комісійні доходи'] + \
                         vnutrishni2.loc['5.10.', 'Внутрішні комісійні витрати']
    clear_vnutri_dohid = pd.DataFrame(clear_vnutri_dohid).T
    fors2 = pd.DataFrame(
            data={
                '№ (рівень 1)': [''],
                'Назва (рівень 1)':['Чистий внутрішній комісійний дохід']
            })
    clear_vnutri_dohid_full = pd.concat([fors2, clear_vnutri_dohid], axis=1)

    zamena_indexa(vnutrishni2)
    data = pd.concat([data, vnutrishni2, clear_vnutri_dohid_full])

    neprocentni = combiend_df_2.iloc[-1].drop(['№ (рівень 2)', 'Назва (рівень 2)', '№ (рівень 3)', 'Назва (рівень 3)',
                                               'Група статей (рівень.4)', 'Наименование группы 4'])
    neprocentni[0] = 'Непроцентні доходи'
    neprocentni = pd.DataFrame(neprocentni).T
    data = pd.concat([data, neprocentni])


    data = data.replace(0, np.nan)

    data['БАНК (%)'] = round(data['БАНК (ф.), грн.'] / data['БАНК (п.), грн.'] * 100)

    data.iloc[:, 4::3] = np.round(data.iloc[:, 3::3].values / data.iloc[:, 2::3].values * 100)
    data = data.replace(np.nan, 0)
    data.iloc[:, 4::3] = data.iloc[:, 4::3].astype(str) + '%'
    data.iloc[:, 3::3] = np.round(data.iloc[:, 3::3] / 1000)
    data.iloc[:, 2::3] = np.round(data.iloc[:, 2::3] / 1000)

    lst_nazva = []
    for i in data.columns[17::3]:
        lst_nazva.append(i)



    for i in data.iloc[:, 4::3]:
        data.rename(columns={i: ' % виконання м/п '}, inplace=True)

    for i in data.iloc[:, 3::3]:
        data.rename(columns={i: '   Факт    '}, inplace=True)

    for i in data.iloc[:, 2::3]:
        data.rename(columns={i: '   План    '}, inplace=True)


    data.set_index('№ (рівень 1)', inplace=True)


    data.to_excel('path_to_file.xlsx')
    # # ========================================================================================================

    wb = openpyxl.load_workbook('path_to_file.xlsx')
    ws = wb['Sheet1']

    ws.insert_cols(18)
    for rows in ws.iter_cols(min_col=18):
        for cell in rows:
            if cell ==18:
                cell.fill = PatternFill(start_color='66B2FF',end_color='66B2FF',fill_type='solid')

    for column in ws.columns:

        lenght = max(len(str(cell.value))for cell in column)
        lenght = lenght if lenght <= 75 else 75
        ws.column_dimensions[column[0].column_letter].width = lenght

    blue = ['Комісійні доходи', 'Комісійні витрати', 'Непроцентні доходи']
    green = ['4.2.1.', '4.2.2.', '4.2.3.', '4.2.4.', '4.2.5.', '4.2.6.', '5.2.1.', '5.2.2.',
             '5.2.3.', '5.2.4.', '5.2.99.', '4.3.', '4.4.', '4.5.', '4.6.', '4.7.', '4.8.', '4.9.']

    tonki = []
    for i in range(1, 15):
        tonki.append('4.2.1.3.' + str(i) + '.')
        tonki.append('4.2.1.4.' + str(i) + '.')


    col_range = ws.max_column

    for i in range(1, len(data.values)+2):
        for col in range(1, col_range + 1):
            cell_header = ws.cell(i, col)
            cell_header.alignment = Alignment(horizontal='right', vertical='center')
            cell_header.font = Font(name='Times New Roman', size=10, bold=True)
            cell_header.border = Border(left=Side(border_style='thin', color='FF000000'),
                                        right=Side(border_style='thin', color='FF000000'),
                                        top=Side(border_style='thin', color='FF000000'),
                                        bottom=Side(style='thin'))
        ws['B'+ str(i)].alignment = Alignment(horizontal='left', vertical='center')
        ws['A'+ str(i)].alignment = Alignment(horizontal='left', vertical='center')
        names = [ws['B' + str(i)].value]
        indexes = [ws['A' + str(i)].value]
        for j in names:
            ws['A' + str(i)].font = Font(name='Times New Roman', size=11, bold=True)
            ws['B' + str(i)].font = Font(name='Times New Roman', size=11, bold=True)
            if j in blue:
                for col in range(1, col_range + 1):
                    cell_header = ws.cell(i, col)
                    cell_header.fill = PatternFill(start_color='66B2FF', fill_type="solid")
                    cell_header.font = Font(name='Times New Roman', size=13,  bold=True)
                ws['B' + str(i)] = ws['B' + str(i)].value.upper()
                ws['B' + str(i)].font = Font(name='Times New Roman', size=13, bold=True)
                ws['A' + str(i)].font = Font(name='Times New Roman', size=13, italic=True, bold=True)
                ws['A' + str(i)].fill = PatternFill(fill_type='solid', start_color='E0FFFF')
                ws['B' + str(i)].fill = PatternFill(fill_type='solid', start_color='66B2FF')

            elif j == 'Чистий комісійний дохід':
                for col in range(1, col_range + 1):
                    cell_header = ws.cell(i, col)
                    cell_header.fill = PatternFill(start_color='5959AB', fill_type="solid")
                    cell_header.font = Font(name='Times New Roman', size=11,  bold=True)
                ws['A' + str(i)].fill = PatternFill(fill_type='solid', start_color='A2CD5A')
                ws['B' + str(i)].font = Font(name='Times New Roman', size=13, bold=True)
                ws['B' + str(i)].fill = PatternFill(fill_type='solid', start_color='5959AB')
        for k in indexes:
            if k in green:
                for col in range(1, col_range + 1):
                    cell_header = ws.cell(i, col)
                    cell_header.fill = PatternFill(start_color='A2CD5A', fill_type="solid")
                    cell_header.font = Font(name='Times New Roman', size=11,  bold=True)
                ws['B' + str(i)] = ws['B' + str(i)].value.upper()
                ws['B' + str(i)].font = Font(name='Times New Roman', size=11, bold=True)
                ws['A' + str(i)].font = Font(name='Times New Roman', size=11, italic=True, bold=True)
                ws['A' + str(i)].fill = PatternFill(fill_type='solid', start_color='BCEE68')
                ws['B' + str(i)].fill = PatternFill(fill_type='solid', start_color='A2CD5A')
            if k == '4.3.':
                wrap_aligmant = Alignment(wrap_text=True)
                ws['B'+str(i)].alignment = wrap_aligmant
            if k == '4.4.':
                wrap_aligmant = Alignment(wrap_text=True)
                ws['B'+str(i)].alignment = wrap_aligmant
            elif k in tonki:
                for col in range(1, col_range + 1):
                    cell_header = ws.cell(i, col)
                    cell_header.font = Font(name='Times New Roman', size=8, italic=True, bold=False)
                ws['B' + str(i)].font = Font(name='Times New Roman', size=8, italic=True, bold=False)
                ws['B' + str(i)].alignment = Alignment(horizontal='right', vertical='center')
                ws['A' + str(i)].font = Font(name='Times New Roman', size=8, italic=True, bold=False)
            if k == '4.2.1.4.9.':
                wrap_aligmant = Alignment(wrap_text=True, horizontal='right', vertical='center')
                ws['B' + str(i)].alignment = wrap_aligmant

    ws.insert_rows(0)

    list_slice = ['РАЗОМ', 'КОРПОРАТИВНИЙ БІЗНЕС', 'МСБ', 'РОЗДРІБНИЙ БІЗНЕС', 'ІНШІ']

    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws['B1'] = 'НЕПРОЦЕНТНІ ДОХОДИ'
    ws['B1'].font = Font(name='Times New Roman', size=16,  bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    rd = ws.row_dimensions[1]
    rs = ws.row_dimensions[2]
    rd.height = 45
    rs.height = 30

    def slice(x, y, z):
        while y != 20:
            ws.merge_cells(start_row=1, start_column=x, end_row=1, end_column=y)
            x += 3
            y += 3
        for i in range(len(list_slice)):
            ws.cell(row=1, column=z).value = list_slice[i]
            ws.cell(row=1, column=z).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=1, column=z).font = Font(name='Times New Roman', size=12, bold=True)
            ws.cell(row=1, column=z).border = Border(left=Side(border_style='thin', color='FF000000'),
                                        right=Side(border_style='thin', color='FF000000'),
                                        top=Side(border_style='thin', color='FF000000'),
                                        bottom=Side(style='thin'))
            z += 3

    slice(3, 5, 3)

    def slice2(x, y, z):
        while y != 60:
            ws.merge_cells(start_row=1, start_column=x, end_row=1, end_column=y)
            x += 3
            y += 3
        for i in range(len(lst_nazva)):
            ws.cell(row=1, column=z).value = lst_nazva[i]
            ws.cell(row=1, column=z).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row=1, column=z).font = Font(name='Times New Roman', size=12, bold=True)
            ws.cell(row=1, column=z).border = Border(left=Side(border_style='thin', color='FF000000'),
                                        right=Side(border_style='thin', color='FF000000'),
                                        top=Side(border_style='thin', color='FF000000'),
                                        bottom=Side(style='thin'))
            z += 3

    slice2(19, 21, 19)




    wb.save('path_to_file.xlsx')

tables(pd.read_excel('Бізнес-юніти накопичувально на 13.12.2021.xlsx'), pd.read_excel('Бізнес накопичувально на 13.12.2021.xlsx'))
