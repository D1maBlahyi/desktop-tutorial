import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
from allocation_func import all_func
from datetime import datetime
from glob import glob
pd.set_option('display.max_rows', 5000)
pd.set_option('display.max_columns', 5000)
pd.set_option('display.width', 5000)
from functools import reduce
from main import person


shablon = pd.read_excel('Шаблон.xlsx')


lst_x = []

a_01 = pd.read_excel('res_01.xlsx')
b_02 = pd.read_excel('res_02.xlsx')
c_03 = pd.read_excel('res_03.xlsx')
a_04 = pd.read_excel('res_04.xlsx')
b_05 = pd.read_excel('res_05.xlsx')
c_06 = pd.read_excel('res_06.xlsx')
# a_07 = pd.read_excel('res_07.xlsx')
# b_08 = pd.read_excel('res_08.xlsx')
# c_09 = pd.read_excel('res_09.xlsx')
# c_10 = pd.read_excel('res_10.xlsx')
# c_11 = pd.read_excel('res_11.xlsx')
# c_12 = pd.read_excel('res_12.xlsx')
def rename_col(df, sym):
    df = df.add_prefix(sym)
    df.rename(columns={f'{sym}Код статьи': 'Код статьи'}, inplace=True)
    return df


a_04 = rename_col(a_04, '_')
b_05 = rename_col(b_05, '!')
c_06 = rename_col(c_06, '&')
# a_07 = rename_col(a_07, '*')
# b_08 = rename_col(b_08, ')')
# c_09 = rename_col(c_09, '(')
# c_10 = rename_col(c_10, '#')
# c_11 = rename_col(c_11, '+')
# c_12 = rename_col(c_12, '$')

for i in range(len(a_01)):
    if a_01['Код статьи'][i] == '!7':
        lst_x.append(list(a_01['Код статьи'][i:]))
lst_x = sum(lst_x, [])




def save_all (file):
    file = file.loc[file['Код статьи'].isin(lst_x)].reset_index(drop=True)
    return file

x_1_all = save_all(a_01)
y_1_all = save_all(b_02)
z_1_all = save_all(c_03)
x_2_all = save_all(a_04)
y_2_all = save_all(b_05)
z_2_all = save_all(c_06)
# x_3_all = save_all(a_07)
# y_3_all = save_all(b_08)
# z_3_all = save_all(c_09)
# z_4_all = save_all(c_10)
# z_5_all = save_all(c_11)
# z_6_all = save_all(c_12)
def dell_all (file):
    file = file[['Код статьи']]
    file = file.loc[~file['Код статьи'].isin(lst_x)].reset_index(drop=True)
    return file

x_1 = dell_all(a_01)
y_1 = dell_all(b_02)
z_1 = dell_all(c_03)
x_2 = dell_all(a_04)
y_2 = dell_all(b_05)
z_2 = dell_all(c_06)
# x_3 = dell_all(a_07)
# y_3 = dell_all(b_08)
# z_3 = dell_all(c_09)
# z_4 = dell_all(c_10)
# z_5 = dell_all(c_11)
# z_6 = dell_all(c_12)


def concat(*args):
    list_dfs = list(args)
    df_concat = pd.concat(list_dfs, join='outer', axis=0)
    return df_concat

concate_tab = concat(x_1[['Код статьи']], y_1[['Код статьи']],  z_1[['Код статьи']], x_2[['Код статьи']],  y_2[['Код статьи']], z_2[['Код статьи']])


# ,
#                      , ,
#                      x_3[['Код статьи']], y_3[['Код статьи']], z_3[['Код статьи']],
#                      z_4[['Код статьи']], z_5[['Код статьи']], z_6[['Код статьи']]

def concat_(*args):
    list_dfs = list(args)
    df_concat = pd.concat(list_dfs, join='outer', axis=1)
    return df_concat

all_rows = pd.DataFrame(data={'all':lst_x})

concat_alloc = concat_(all_rows, x_1_all.iloc[:, 1:], y_1_all.iloc[:, 1:],  z_1_all.iloc[:, 1:], x_2_all.iloc[:, 1:], y_2_all.iloc[:, 1:], z_2_all.iloc[:, 1:])

# ,
#                        z_2_all.iloc[:, 1:],
#                      x_3_all.iloc[:, 1:], y_3_all.iloc[:, 1:], z_3_all.iloc[:, 1:], z_4_all.iloc[:, 1:], z_5_all.iloc[:, 1:],
#                        z_6_all.iloc[:, 1:]


concate_tab = concate_tab.drop_duplicates().reset_index(drop=True)



CHek = shablon['Код статьи'][shablon['Чек'] != 1.0].to_list()
chek_2 = shablon['Код статьи'][shablon['Чек_2'] == 1.0].to_list()
chek_3 = shablon['Код статьи'][shablon['Чек_3'] == 1.0].to_list()

df_lst_map = []
for i in range(len(CHek)):
    second_df = []
    if CHek[i] not in chek_2:
        df_lst_map.append(CHek[i])
    elif CHek[i] in chek_2:
        for j in range(len(shablon)):
            if shablon['Код статьи'][j].startswith(CHek[i]):
                if shablon['Код статьи'][j] not in chek_3:
                    second_df.append(shablon['Код статьи'][j])
        for k in range(len(concate_tab)):
            if str(concate_tab['Код статьи'][k]).startswith(CHek[i]):
                second_df.append(concate_tab['Код статьи'][k])
        second_df = sorted(second_df)
        for h in range(len(second_df)):
            df_lst_map.append(second_df[h])

# Полный порядок только статтей
data = pd.DataFrame(data={'Код статьи': df_lst_map})

lst_4 = []
for i in range(len(data)):
    if data['Код статьи'][i] == '4.8.9.25.':
        lst_4.append(i)
data.drop(lst_4[0], axis=0, inplace=True)

data = data.drop_duplicates(keep='first').reset_index(drop=True)




def nlo(state, data, chek_2):
    lst_key = []
    lst_val = []
    for i in range(len(data)):
        if data['Код статьи'][i].startswith(state):
            lst_key.append(i)
            lst_val.append(data['Код статьи'][i])
    if len(lst_key) > 2:
        while len(lst_key) > 2:
            data = data.drop(lst_key[-1], axis=0)
            line = pd.DataFrame({"Код статьи": lst_val[-1]}, index=[lst_key[1] + 0.5])
            data = data.append(line, ignore_index=False)
            data = data.sort_index().reset_index(drop=True)
            chek_2.append(lst_val[-1])
            lst_key.pop(-1)
            lst_val.pop(-1)

    return data


data = nlo('5.3.6.38.', data, chek_2)
data = nlo('5.4.6.5.26.', data, chek_2)


lst_rez = []
lst_rez_drop = []

for i in range(len(data)):
    if data['Код статьи'][i] == '5.5.7.2.1.2.' or data['Код статьи'][i] == '5.5.7.2.2.2.':
        data = data.drop(i, axis=0)
data = data.reset_index(drop=True)

for i in range(len(data)):
    for j in range(4, 8):
        if data['Код статьи'][i] == f'5.5.7.{j}.':
            lst_rez.append(data['Код статьи'][i])
            lst_rez_drop.append(i)


for i in range(len(lst_rez_drop)):
    data = data.drop(lst_rez_drop[i], axis=0)
data = data.reset_index(drop=True)

line = 0
for i in range(len(data)):
    if data['Код статьи'][i] == '5.5.7.':
        line = pd.DataFrame({"Код статьи": '5.5.7.2.1.2.'}, index=[i + 0.5])
        data = data.append(line, ignore_index=False)
        data = data.sort_index().reset_index(drop=True)

for i in range(len(data)):
    if data['Код статьи'][i] == '5.5.7.2.1.2.':
        line = pd.DataFrame({"Код статьи": '5.5.7.2.2.2.'}, index=[i + 0.5])
        data = data.append(line, ignore_index=False)
        data = data.sort_index().reset_index(drop=True)

lst_rez = lst_rez[::-1]

for i in range(len(data)):
    if data['Код статьи'][i] == "#16":
        for j in range(len(lst_rez)):
            line = pd.DataFrame({"Код статьи": lst_rez[j]}, index=[i + 0.5])
            data = data.append(line, ignore_index=False)
            data = data.sort_index().reset_index(drop=True)


data = data.drop_duplicates().reset_index(drop=True)

# all_rows = pd.DataFrame(data={'all':lst_x})
#
# data = pd.concat([data, all_rows], axis=0)
# data = data['Код статьи'].fillna(data['all']).reset_index(drop=True)



def merge(*args):
    list_dfs = list(args)
    df_merged = reduce(lambda left, right: pd.merge(left, right, on=['Код статьи'],
                                                    how='left'), list_dfs)
    return df_merged


finish_data_1 = merge(data, a_01, b_02, c_03, a_04, b_05, c_06)
# , , , a_07, b_08, c_09, c_10, c_11, c_12


save_col = finish_data_1.columns
finish_data_1.columns = list(range(1, len(finish_data_1.columns)+1))
concat_alloc.columns = list(range(1, len(finish_data_1.columns)+1))
res_f = pd.concat([finish_data_1, concat_alloc], axis=0).reset_index(drop=True)
res_f.columns = save_col



finish_data_1 = res_f


def culc(x):
    finish_data_1.iloc[:, x] = finish_data_1.iloc[:, x::29].sum(axis=1)
    return finish_data_1
for i in range(2, 30):
    culc(i)


for i in range(len(finish_data_1)):
    for j in range(1, len(finish_data_1.columns)):
        if type(finish_data_1.iloc[i, 1]) != str:
            if type(finish_data_1.iloc[i, j]) == str:
                finish_data_1.iloc[i, 1] = finish_data_1.iloc[i, j]

finish_data_1 = finish_data_1.iloc[:, :30]


res = finish_data_1

num_col = 2
while num_col <= 27:
    for i in range(len(res)):
        res.iloc[i, num_col + 2] = res.iloc[i, num_col + 1] - res.iloc[i, num_col]
        if -1 < res.iloc[i, num_col] < 1:
            res.iloc[i, num_col] = 0
        if res.iloc[i, num_col] == 0:
            if res.iloc[i, num_col + 1] == 0:
                res.iloc[i, num_col + 3] = '-'
            elif res.iloc[i, num_col + 1] > 0:
                res.iloc[i, num_col + 3] = '300%+'
            elif res.iloc[i, num_col + 1] < 0:
                res.iloc[i, num_col + 3] = '- 300%+'
        elif res.iloc[i, num_col + 1] == 0:
            if res.iloc[i, num_col] > 0:
                res.iloc[i, num_col + 3] = '0%'
            elif res.iloc[i, num_col] < 0:
                res.iloc[i, num_col + 3] = '300%+'
        elif res.iloc[i, num_col + 1] < 0 and res.iloc[i, num_col] < 0:
            res.iloc[i, num_col + 3] = res.iloc[i, num_col] / res.iloc[i, num_col + 1]
            if res.iloc[i, num_col + 3] > 3:
                res.iloc[i, num_col + 3] = '300%+'
        elif res.iloc[i, num_col + 1] > 0 and res.iloc[i, num_col] > 0:
            res.iloc[i, num_col + 3] = res.iloc[i, num_col + 1] / res.iloc[i, num_col]
            if res.iloc[i, num_col + 3] > 3:
                res.iloc[i, num_col + 3] = '300%+'
        else:
            res.iloc[i, num_col + 3] = '300%+'
    num_col += 4

wb = Workbook()
ws = wb.active

for r in dataframe_to_rows(res, index=False):
    ws.append(r)

for column in ws.columns:
    lenght = max(len(str(cell.value)) for cell in column)
    ws.column_dimensions[column[0].column_letter].width = lenght

union = chek_3 + ['#18']
blue = []
grey = chek_2
thin = []

all_finish = pd.read_excel('res_all_.xlsx')
all_color = []
for i in range(len(all_finish)):
    if len(all_finish.iloc[i, 1]) > 4:
        all_color.append(all_finish.iloc[i, 1])


for i in range(len(union)):
    if '#' in union[i]:
        blue.append(union[i])
    elif '4.4.' in union[i]:
        blue.append(union[i])
    elif '4.14.' in union[i]:
        blue.append(union[i])
    elif '4.5.' in union[i]:
        blue.append(union[i])
for word in blue:
    union.remove(word)

for i in range(len(res)):
    if res['Код статьи'][i] not in union and res['Код статьи'][i] not in grey and res['Код статьи'][i] not in blue:
        thin.append(res['Код статьи'][i])

all_col = ['1.2.', '1.7.', '1.1.']
thin.remove('!7')
thin.remove('1.2.')
thin.remove('1.7.')
thin.remove('1.1.')


lists = [[] for _ in range(12)]
list_all_gr = []
list_alloc_2 = []
col_range = ws.max_column

for i in range(1, len(res.values) + 2):

    for col in range(1, col_range + 1):
        cell_header = ws.cell(i, col)
        cell_header.alignment = Alignment(horizontal='right', vertical='center')
        cell_header.font = Font(name='Times New Roman', size=10, bold=True)
        cell_header.border = Border(left=Side(border_style='thin', color='FF000000'),
                                    right=Side(border_style='thin', color='FF000000'),
                                    top=Side(border_style='thin', color='FF000000'),
                                    bottom=Side(style='thin'))
    ws['B' + str(i)].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws['A' + str(i)].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)


    names = [ws['B' + str(i)].value]
    indexes = [ws['A' + str(i)].value]

    def main_format(color, size):
        for col in range(1, col_range + 1):
            cell_header = ws.cell(i, col)
            cell_header.fill = PatternFill(start_color=color, fill_type="solid")
            cell_header.font = Font(name='Times New Roman', size=size, bold=True)
            cell_header.number_format = '#,##0'
        ws['B' + str(i)] = ws['B' + str(i)].value.upper()
        ws['B' + str(i)].font = Font(name='Times New Roman', size=size, bold=True)
        ws['A' + str(i)].font = Font(name='Times New Roman', size=size, italic=True, bold=True)
        ws['B' + str(i)].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws['A' + str(i)].fill = PatternFill(fill_type='solid', start_color=color)
        ws['B' + str(i)].fill = PatternFill(fill_type='solid', start_color=color)
        for col in range(6, col_range + 1, 4):
            cell_header = ws.cell(i, col)
            cell_header.number_format = '0%'
    def low_format(color):
        for col in range(1, col_range + 1):
            cell_header = ws.cell(i, col)
            cell_header.fill = PatternFill(start_color=color, fill_type="solid")
            cell_header.font = Font(name='Times New Roman', size=9, italic=True, bold=False)
            cell_header.number_format = '#,##0'
        ws['B' + str(i)].font = Font(name='Times New Roman', size=9, bold=False)
        ws['B' + str(i)].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws['A' + str(i)].font = Font(name='Times New Roman', size=9, bold=False)
        ws['A' + str(i)].fill = PatternFill(fill_type='solid', start_color=color)
        ws['B' + str(i)].fill = PatternFill(fill_type='solid', start_color=color)
        for col in range(6, col_range + 1, 4):
            cell_header = ws.cell(i, col)
            cell_header.number_format = '0%'

    for j in indexes:
        if j in blue:
            main_format('B0E0E6', 12)
            lists[0].append(i)
        elif j in union:
            main_format('3CB371', 11)
            lists[1].append(i)
        elif j in grey:
            main_format('C0C0C0', 10)
            lists[2].append(i)
        elif j == '!7':
            main_format('FFB233', 12)
        elif len(j.split('.')[:-1]) == 4:
            low_format('BD5E86')
            lists[3].append(i)
        elif len(j.split('.')[:-1]) == 5:
            low_format('DDE533')
            lists[4].append(i)
        elif len(j.split('.')[:-1]) == 6:
            low_format('90A8B8')
            lists[5].append(i)
        elif len(j.split('.')[:-1]) == 7:
            low_format('D1C1E2')
            lists[6].append(i)
        elif len(j.split('.')[:-1]) == 8:
            low_format('E1B0A2')
            lists[7].append(i)
        elif len(j.split('.')[:-1]) == 9:
            low_format('A44953')
            lists[8].append(i)
        elif len(j.split('.')[:-1]) == 10:
            low_format('5D8668')
            lists[9].append(i)
        elif j in all_col:
            main_format('E5B9EB', 11)
            list_all_gr.append(i)
        elif j in all_color:
            low_format('FFFFFF')
            list_alloc_2.append(i)

for i in range(len(lists[0])):
    lists[0][i] += 2
for i in range(0, len(lists[0]) - 1, 1):
    ws.row_dimensions.group(lists[0][i] + 1, lists[0][i + 1] - 1, hidden=True, outline_level=1)
for i in range(len(list_all_gr)):
    list_all_gr[i] += 2
for i in range(0, len(list_all_gr) - 1, 1):
    ws.row_dimensions.group(list_all_gr[i] + 1, list_all_gr[i + 1] - 1, hidden=True, outline_level=3)
ws.row_dimensions.group(list_all_gr[-1] + 1, lists[0][-1] - 1, hidden=True, outline_level=3)

def dimension(lst, lvl, step, lst_2):
    for i in range(len(lst)):
        lst[i] += 2
    for i in range(0, len(lst)-1, step):
        for j in range(1, len(lst_2)-1):
            if lst[i] < lst[i+1] < lst_2[j]:
                ws.row_dimensions.group(lst[i]+1, lst[i+1]-1, hidden=True, outline_level=lvl)
                break
            elif lst[i] < lst_2[j] < lst[i + 1]:
                ws.row_dimensions.group(lst[i] + 1, lst_2[j] - 1, hidden=True, outline_level=lvl)
                break

dimension(lists[1], 2, 1, lists[0])
lists[1] = list(sorted(lists[0] + lists[1]))

dimension(lists[2], 3, 1, lists[1])
lists[2] = list(sorted(lists[1] + lists[2]))

dimension(lists[3], 4, 1, lists[2])
lists[3] = list(sorted(lists[2] + lists[3]))

dimension(lists[4], 5, 1, lists[3])



for i in range(1, len(res.values) + 2):
    indexes = [ws['A' + str(i)].value]
    for j in indexes:
        if '#' in j or '!' in j:
            ws['A' + str(i)] = np.nan

ws.insert_rows(0)
ws.insert_rows(0)

for i in range(3, len(res.columns)):
    ws.cell(row=3, column=i).alignment = Alignment(horizontal='center', vertical='center')

lst_names = ['Консолідовано Банк', 'Корпоративний бізнес', 'Корпоратив - мережа', 'VIP Корпоратив',
             'Малий та середній бізнес', 'Роздрібний  банкінг', 'Інші']

ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)

for i in range(2, 4):
    ws['A2'] = 'Код статей'
    ws['A2'].font = Font(name='Times New Roman', size=14, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A' + str(i)].border = Border(left=Side(border_style='thick', color='FF000000'),
                                              right=Side(border_style='thick', color='FF000000'),
                                              top=Side(border_style='thick', color='FF000000'),
                                              bottom=Side(style='thin', color='FF000000'))
    ws['B2'] = 'Найменування статей'
    ws['B2'].font = Font(name='Times New Roman', size=14, bold=True)
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B' + str(i)].border = Border(left=Side(border_style='thick', color='FF000000'),
                                     right=Side(border_style='thick', color='FF000000'),
                                     top=Side(border_style='thick', color='FF000000'),
                                     bottom=Side(style='thin', color='FF000000'))
    rd = ws.row_dimensions[2]
    rs = ws.row_dimensions[3]
    rd.height = 35
    rs.height = 25
    rq = ws.row_dimensions[1]
    rq.height = 25

for i in range(3, len(res.columns)):
    ws.cell(row=2, column=i).border = Border(left=Side(border_style='thick', color='FF000000'),
                                             right=Side(border_style='thick', color='FF000000'),
                                             top=Side(border_style='thick', color='FF000000'),
                                             bottom=Side(style='thick', color='FF000000'))

def slice(x, y, z, limit, lst):
    while y <= limit:
        ws.merge_cells(start_row=2, start_column=x, end_row=2, end_column=y)
        x += 4
        y += 4
    for i in range(len(lst)):
        ws.cell(row=2, column=z).value = lst[i]
        ws.cell(row=2, column=z).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=2, column=z).font = Font(name='Times New Roman', size=12, bold=True)

        z += 4


slice(3, 6, 3, len(res.columns), lst_names)

for column in ws.columns:
    lenght = max(len(str(cell.value)) for cell in column)
    if lenght > 75:
        lenght = 100
    ws.column_dimensions[column[0].column_letter].width = lenght

freez_obl = ws['C4']
ws.freeze_panes = freez_obl
wb.save(f"C:\\Users\\admin\PycharmProjects\\fin_b\\files\\Накопичувально_2023.xlsx")
wb.close()
res.to_excel(f'res_nak.xlsx', index=False)

